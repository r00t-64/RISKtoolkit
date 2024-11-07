from sqlalchemy import create_engine, text,types
from datetime import datetime, timedelta
import re,os,json
import pandas as pd
import numpy as np
from exchangelib import Credentials, Account, Message, FileAttachment,HTMLBody
from openpyxl.styles import Font, PatternFill, Border, Alignment, numbers

from riesgosutils.oracle_connection import OracleConnection
from riesgosutils.styler_html import StylerHTML

import win32com.client
outlook = win32com.client.Dispatch('outlook.application')

""" 
def db_select(cursor, query, ): # takes the request text and sends it to the data_frame
    response = {"is_ok": True, "error": "", "content":None}

    try:
        if query.lower().endswith(".sql"):
            __sql_str = parse_query_file(query)
        else:
            if query.lower().startswith("select"):
                __sql_str = query
            else:
                __sql_str = 'select * from ' + query
        
        cursor.execute(__sql_str)
        __col = [column[0].lower() for column in cursor.description]
        __df = pd.DataFrame.from_records(cursor, columns=__col) 
        
        if len(__df) > 0:
            response["content"] = __df
        else:
            response["content"] = None
        

    except Exception as e:
        response = {"is_ok": False, "error": str(e), "content": None}

    finally:
        return response 

    def execute_query_from_file(cursor, file_sql, parameters, output_type='default'):
        with open(file_sql, 'r', encoding='latin-1') as __file:
            __query_template = __file.read()

        __query_str = __query_template.format(**parameters)
        cursor.execute(__query_str)
        __result = cursor.fetchall()

        if __result:
            if output_type == 'dataframe':
                __col_names = [column[0].lower() for column in cursor.description]
                __df = pd.DataFrame(__result, columns=__col_names)
                return __df  
            else:
                __result_list = [value for row in __result for value in row]
                return __result_list 
        else:
            return []  

"""
def funcion_sql(cursor, query, parameters=None, output_type='default'):
    response = {"is_ok": True, "error": ""}
    output = None

    try:
        # Cargar SQL desde un archivo si el nombre del archivo termina en .sql
        if query.lower().endswith(".sql"):
            with open(query, 'r', encoding='latin-1') as file:
                query_template = file.read()
            # Formatear la consulta con los parámetros proporcionados
            query_str = query_template.format(**parameters)
        else:
            # Determinar la cadena de consulta según el input
            if len(query.split()) == 1:
                # Si la consulta es una sola palabra, se asume que es el nombre de una tabla
                query_str = 'SELECT * FROM ' + query
            else:
                # Si la consulta comienza con "select", se usa directamente
                # De lo contrario, se considera una consulta válida completa
                query_str = query if query.lower().startswith("select") else query

        # Ejecutar la consulta SQL
        cursor.execute(query_str)
        result = cursor.fetchall()
        # Obtener los nombres de las columnas
        col_names = [column[0].lower() for column in cursor.description]

        # Procesar el resultado según el tipo de salida solicitado
        if output_type == 'dataframe':
            # Crear un DataFrame, incluso si no hay filas (solo columnas)
            output = pd.DataFrame(result, columns=col_names) if result else pd.DataFrame(columns=col_names)
        elif output_type == 'json':
            # Convertir el resultado a un diccionario y luego a formato JSON
            result_dict = [dict(zip(col_names, row)) for row in result]
            output = json.dumps(result_dict, ensure_ascii=False)
        else:
            # Devolver nombres de columnas más las filas como lista de listas
            output = [col_names] + [list(row) for row in result] if result else [col_names]

    except Exception as e:
        # En caso de error, actualizar el mensaje en la respuesta
        response = {"is_ok": False, "error": str(e)}

    finally:
        return output, response

def proceso_sql(engine, query, parameters=None):
    """
    Ejecuta sentencias SQL desde un archivo (.sql) o desde una cadena directa, manejando parámetros y control de errores.
    :param engine: Motor SQLAlchemy para ejecutar las consultas.
    :param query: Ruta al archivo SQL o cadena de consulta SQL directa.
    :param parameters: Diccionario de parámetros para formatear las consultas SQL.
    :param sep: Separador para dividir múltiples consultas (por defecto ';').
    :return: Diccionario con el estado de la ejecución.
    """
    __response = {"is_ok": True, "error": ""}
    
    try:
        if query.lower().endswith(".sql"):

            filename=query
            with open(filename,encoding='utf-8-sig') as file:
                statements = re.split(';', file.read(), flags=re.MULTILINE)
                for statement in statements:
                    try:
                        engine.execute(text(statement))
                    except Exception as __e:
                        print(f"Database error occurred: {str(__e)}")  
                        continue
        else:
            engine.execute(text(query))

    except Exception as __e:
        __response = {"is_ok": False, "error": str(__e)}
    
    return __response


def proceso_sql_string(engine, query_str: str ):
    __response = {"is_ok": True, "error": ""}
    try:
        engine.execute(text(query_str))
    except Exception as __e:
        __response = {"is_ok": False, "error": str(__e)}

    return __response



def test_datalake(cursor, bc_today_date: str):
    try:
        __date_obj = datetime.strptime(bc_today_date, "%d/%m/%Y")
    except ValueError as e:
        raise ValueError(f"Invalid date format for bc_today_date: {bc_today_date}. Expected format: 'dd/mm/YYYY'. Error: {e}")

    __datalake_date = __date_obj.strftime("%Y%m%d")
    
    try:
        __response = funcion_sql(cursor, f'datalake.lkta_credito_vigente_{__datalake_date}@bdbi fetch first 1 rows only' , output_type='dataframe')
    except Exception as e:
        raise RuntimeError(f"Database query execution failed: {str(e)}")

    return __response.get("is_ok", False)
    
def get_last_day_of_month(date):
    __first_day_of_current_month = date.replace(day=1)
    __last_day_of_previous_month = __first_day_of_current_month - timedelta(days=1)
    return __last_day_of_previous_month


def send_mail(subject, body, to_email, cc_email=None, attachment_paths=None, outlookacc=None, password=None, draft=False):
    try:
    # Verifica individualmente si alguno de los parámetros obligatorios es None
        if subject is None:
            raise ValueError("El parámetro 'subject' es obligatorio y no puede ser None.")
        if body is None:
            raise ValueError("El parámetro 'body' es obligatorio y no puede ser None.")
        if to_email is None:
            raise ValueError("El parámetro 'to_email' es obligatorio y no puede ser None.")
        if outlookacc is None:
            raise ValueError("El parámetro 'outlookacc' es obligatorio y no puede ser None.")
        if password is None:
            raise ValueError("El parámetro 'password' es obligatorio y no puede ser None.")
        
        
        __credentials = Credentials(username=outlookacc, password=password)
        __account = Account(primary_smtp_address=outlookacc, credentials=__credentials, autodiscover=True)

        __email = Message(
            account=__account,
            subject=subject,
            body=HTMLBody(body)
        )
        
        if attachment_paths:
            for __attachment_path in attachment_paths:
                if __attachment_path:  # Verifica que no sea None o vacío
                    with open(__attachment_path, 'rb') as __file:
                        __file_name = os.path.basename(__attachment_path)
                        __file_content = __file.read()
                        __file_attachment = FileAttachment(name=__file_name, content=__file_content)
                        __email.attach(__file_attachment)

        __email.to_recipients = to_email if to_email else []
        __email.cc_recipients = cc_email if cc_email else []

        if draft:
            __email.folder = __account.drafts
            __email.save()
            print("Correo guardado en 'bandeja de salida'")
        else:
            __email.send()
            print("Correo enviado correctamente")
    except Exception as e:
        print(f"Error al enviar correo\n", str(e))

    

def insert_dataframe(engine, df, table_name, schema= "", index=False, if_exists='append'):
    __response = {"is_ok": True, "error": ""}

    try:
        __column_types = {}
        for __col, __dtype in df.dtypes.items():
            if __dtype == 'int64' or __dtype == 'int32':
                __column_types[__col] = types.Integer
            elif __dtype == 'object':
                __max_length = df[__col].str.len().max()
                __column_types[__col] = types.VARCHAR(length=max(__max_length, 100))
            elif __dtype == 'float64':
                __column_types[__col] = types.Float
            elif __dtype == 'datetime64[ns]':
                __column_types[__col] = types.Date

        if not index:
            df = df.reset_index(drop=True)
        df.to_sql(con = engine, schema= schema, name= table_name, dtype=__column_types, if_exists=if_exists, index=False, chunksize=1000)
    
    except Exception as __e:
        __response = {"is_ok": False, "error": str(__e)}

    return __response

def split_sheet_by_categories(input_file_name, output_file_name, category):
    """
    Divide una hoja de Excel en múltiples hojas basadas en una columna de categoría especificada.

    Parámetros:
    - input_file_name (str): Nombre del archivo Excel de entrada.
    - output_file_name (str): Nombre del archivo Excel de salida.
    - category (str): Nombre de la columna de categoría por la cual dividir.
    """
    excel_input_filename = input_file_name
    excel_output_filename = output_file_name
    total_sheet_name = "TOTAL"
    pivot_sheet_name = category
    table_style = "Table Style Light 9"

    df = pd.read_excel(excel_input_filename, engine="openpyxl")
    columns = [{"header": column} for column in df.columns]

    categories = df[pivot_sheet_name].unique()
    categories = np.sort(categories)

    writer = pd.ExcelWriter(excel_output_filename, engine="xlsxwriter")

    df.to_excel(writer, sheet_name=total_sheet_name, index=False)
    ws_total_table = writer.sheets[total_sheet_name]
    (max_row, max_col) = df.shape
    ws_total_table.add_table(
        0, 0, max_row, max_col - 1, {"columns": columns, "style": table_style}
    )

    for category in categories:
        filtered_table = df[df[pivot_sheet_name] == category]
        filtered_table.to_excel(writer, sheet_name=str(category), index=False)
        ws_filtered_table = writer.sheets[str(category)]
        (max_row, max_col) = filtered_table.shape
        ws_filtered_table.add_table(
            0, 0, max_row, max_col - 1, {"columns": columns, "style": table_style}
        )

    writer.save()


def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        # Copy font style
        target_cell.font = Font(name=source_cell.font.name, size=source_cell.font.size, bold=source_cell.font.bold,
                                italic=source_cell.font.italic, vertAlign=source_cell.font.vertAlign,
                                underline=source_cell.font.underline, strike=source_cell.font.strike,
                                color=source_cell.font.color)
        
        # Copy fill style
        if source_cell.fill.fill_type != "none":
            target_cell.fill = PatternFill(fill_type=source_cell.fill.fill_type, start_color=source_cell.fill.start_color,
                                           end_color=source_cell.fill.end_color)
        
        # Copy border style
        target_cell.border = Border(left=source_cell.border.left, right=source_cell.border.right, top=source_cell.border.top,
                                    bottom=source_cell.border.bottom, diagonal=source_cell.border.diagonal,
                                    diagonal_direction=source_cell.border.diagonal_direction,
                                    outline=source_cell.border.outline, vertical=source_cell.border.vertical,
                                    horizontal=source_cell.border.horizontal)
        
        # Copy alignment style
        target_cell.alignment = Alignment(horizontal=source_cell.alignment.horizontal, vertical=source_cell.alignment.vertical,
                                          text_rotation=source_cell.alignment.text_rotation, wrap_text=source_cell.alignment.wrap_text,
                                          shrink_to_fit=source_cell.alignment.shrink_to_fit, indent=source_cell.alignment.indent)
        
        # Copy number format style
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        
        # Copy string format (if it's a string cell)
        if source_cell.data_type == "s":
            target_cell.value = source_cell.value
            target_cell.data_type = "s"

def copy_column_style(source_sheet, source_column, target_sheet, target_column):
    for row in range(1, source_sheet.max_row + 1):
        source_cell = source_sheet.cell(row=row, column=source_column)
        target_cell = target_sheet.cell(row=row, column=target_column)
        copy_cell_style(source_cell, target_cell)

def days_in_month(date_x):
    try:
        if type(date_x).__name__ == 'str':
            # Parse the input date string to a datetime object
            __now = datetime.strptime(date_x, '%d/%m/%Y')

        # Get the first day of the current month
        first_day_of_month = datetime(__now.year, __now.month, 1)

        # Calculate the first day of the next month
        if __now.month == 12:
            first_day_of_next_month = datetime(__now.year + 1, 1, 1)
        else:
            first_day_of_next_month = datetime(__now.year, __now.month + 1, 1)

        # Calculate the number of days in the current month
        days_in_month = (first_day_of_next_month - first_day_of_month).days

        return days_in_month
    except Exception:
        return 0


# def parse_query_file(query):
#     try: 
#         with open(query, encoding='latin-1') as sql_file: 
#             sql_str = sql_file.read()
#     except Exception as e:
#         sql_str = ""
#     finally:
#         return sql_str